class Attendance:

    def __init__(self, attendance_file, user_data_file):
        self.attendanceFile = attendance_file
        self.userFile = user_data_file
        self.entries = []
        self.emp_mapping = dict()
        self.user_data = dict()
        self.date = None
        self.invalids = []

    def parse_input(self):

        def format_date(date):
            from datetime import datetime

            date_obj = datetime.strptime(date, '%Y-%m-%d')
            return date_obj.strftime('%d %a')
        
        for line in self.attendanceFile.getvalue().decode('utf-8').splitlines():
                emp_id, date, time = line.split()[:3]
                self.date = date
                emp_id = emp_id.strip()
                date = format_date(date.strip())
                time = time.strip()

                self.entries.append((emp_id, date, time))
        
    def init_dates(self, emp):

        date = self.entries[0][1]
        year, month, date = self.date.split('-')

        import calendar
        from datetime import datetime

        num_days = calendar.monthrange(int(year), int(month))[1]

        for i in range(1, num_days + 1):
            date_obj = datetime(int(year), int(month) , i)
            date = date_obj.strftime('%d %a')
            emp[date] = {
                'entry': None,
                'exit': None
            }
        
        return emp

    def parse_entries(self):

        for entry in self.entries:
            emp_id, date, time = entry

            emp_id = int(emp_id)
            if emp_id not in self.user_data:
                self.user_data[emp_id] = self.init_dates(dict())

            # entry for first time
            if self.user_data[emp_id][date]['entry'] is None:
                self.user_data[emp_id][date]['entry'] = time
                
            # exit for second time
            else:
                self.user_data[emp_id][date]['exit'] = time
        
        # sort self.user_data by key
        self.user_data = dict(sorted(self.user_data.items()))
        
    
    def print_data(self):
        for emp_id, emp_data in self.user_data.items():
            print(f"Employee ID: {emp_id}")
            for date, data in emp_data.items():
                print(f"\t{date}: Entry: {data['entry']} Exit: {data['exit']}")

    def map_employees(self):
        for line in self.userFile.getvalue().decode('utf-8').splitlines():
                details = line.split()
                emp_id = int(details[-1].strip())
                details.pop()
                emp_name = " ".join(details)

                self.emp_mapping[emp_id] = emp_name

        # sort emp_mapping by emp_id
        self.emp_mapping = dict(sorted(self.emp_mapping.items()))
        
    def write_excel(self):
        # write "Monthly Status Report (Basic Work Duration)" using openpyxl

        import openpyxl
        from openpyxl.styles import Alignment

        wb = openpyxl.Workbook()
        ws = wb.active

        # merge columns of 1st row, insert heading, make bold, center align
        year, month, date = self.get_date()

        ws.merge_cells('A1:X1')
        ws['A1'] = f'{month} {year} Status Report (Basic Work Duration)'
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = openpyxl.styles.Font(bold=True)

        col = 2
        row = 3

        # insert days
        sample_data = self.user_data[list(self.user_data.keys())[0]]
        ws.cell(row=row, column=1, value='Days')
        for date, data in sample_data.items():
            ws.cell(row=row, column=col, value=date)
            col += 1
            
        row = 6

        for id, emp_data in self.user_data.items():

            if(self.emp_mapping.get(id) == None):
                self.invalids.append(id)
                continue

            col = 1
           
            # written employee id and name
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + len(emp_data) - 1)
            ws.cell(row=row, column=col, value=f"Employee ID: {id} \t Employee Name: {self.emp_mapping[id]}")
            ws.cell(row=row, column=col).font = openpyxl.styles.Font(bold=True)

            row += 1

            # written entry and exit time
            ws.cell(row=row, column=col, value='In Time')
            ws.cell(row=row + 1, column=col, value='Out Time')
            ws.cell(row=row + 2, column=col, value='Duration')
            
            col+=1
            for date, data in emp_data.items():
                ws.cell(row=row, column=col, value=data['entry'])
                ws.cell(row=row + 1, column=col, value=data['exit'])

                # difference between entry and exit time
                if(data['entry'] is not None and data['exit'] is not None):
                    ws.cell(row = row + 2, column=col, value=f"{self.time_difference(data['entry'], data['exit'])}")
                col += 1
            
            row += 4
        
        
        return wb, f'Monthly Status Report {month} {year}.xlsx'

    def time_difference(self, entry, exit):
        from datetime import datetime

        entry = datetime.strptime(entry, '%H:%M:%S')
        exit = datetime.strptime(exit, '%H:%M:%S')

        return exit - entry
    
    def get_date(self):
        year, month, date = self.date.split('-')
        
        # Month to be displayed in words
        import calendar
        month = calendar.month_name[int(month)]
        return [year, month, date]
    
def main():
    attendance = Attendance()
    attendance.parse_input()
    attendance.parse_entries()
    attendance.map_employees()
    attendance.write_excel()
